# Revision History
#
#
# Authors: Elvis Segbeaya
# Dealing with WellData in general. Getting File ready for different Use Case Scenarios
#

import csv
import logging
import time
from datetime import datetime as dt, date, datetime, timedelta
import pandas as pd
import SampleHelper
import welldataAPI
from retry import retry
from tenacity import retry, stop_after_attempt, wait_fixed
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl import load_workbook
from openpyxl.formatting.rule import IconSet, FormatObject, Rule, CellIsRule, IconSetRule
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import PatternFill, Font, Border
import re
import numpy as np
from pydantic import BaseModel
from pydantic.dataclasses import dataclass
KeepCacheForDevelopment = False


# creating a logging file

# come here to edit the attributes that you want to hold.
# get jobID, rigNumber, WellName, Startdate, EndDate #adding owner since we should verify who the owner is

class Well(BaseModel):
    jobID = ""
    jobName = ""
    Owner = ""
    rigNumber = ""
    Startdate = ""
    EndDate = ""



class UnitV1(BaseModel):
    id: str
    name: str
    abbreviation: str


class Attribute(BaseModel):
    id: str
    mode: str


def thresholdCheck(min=0, max=150000, input=1):
    if input > min and input < max:
        print(f'{input} is within range')
        return True
    else:
        print(f'{input} is not within range of {min} - {max}')
        return False


def main():
    #######################################################################
    # Setup
    #######################################################################

    # SampleHelper file is used to setup the configuration
    SampleHelper.SetupLogging(logging, logging.DEBUG)
    SampleHelper.SetupLocale()

    SampleHelper.VersionCheck()

    # WellData Config
    configFile = SampleHelper.SetConfigFile(welldataAPI.defaultConfig())

    CFGdefault = SampleHelper.readConfig(configFile, welldataAPI.defaultConfig())

    # parses the api request -> this currently uses the welldata.cfg as the config file
    CFG = SampleHelper.readConfig(configFile, welldataAPI.serverConfig('welldata net'))
    CFG.update(CFGdefault)

    # Process Name, used for stateless process control
    today = date.today().strftime('%m-%d-%Y')
    today2 = date.today().strftime('%Y-%m-%d')
    datetime_string_from = today2 + 'T06:05:17'
    datetime_string_to = today2 + 'T06:06:17'
    print(today)
    if CFG['FromHours'] != '':
        datetime_string_from = CFG['FromHours']

    if CFG['ToHours'] != '':
        datetime_string_to = CFG['ToHours']


    print(datetime_string_from)
    print(datetime_string_to)
    processName = 'EDR Reports'

    DataSource = None
    if CFG['APIUrl'] == 'https://data.welldata.net/api/v1':
        DataSource = 'WellData Data API .NET'
    else:
        logging.error("Unknown Datasource {}. Aborting".format(DataSource))
        quit()

    # Original Urls
    # URLs = welldataAPI.URLs(CFG['APIUrl'], CFG['ContractorName'], CFG['OperatorName'], CFG['SpudYearStart'], CFG['SpudYearEnd'], CFG['JobStatus'])
    URLs_v1 = welldataAPI.URLs_v1(CFG['APIUrl'], CFG['ContractorName'], CFG['OperatorName'], CFG['JobStatus'])

    # Generates Token for Header
    token = welldataAPI.getToken(CFG['APIUrl'], CFG['appID'], CFG['username'], CFG['password'])
    # ONLY Get Patterson wells - getWells will filter them for us



    # # Writing DataFrame to Excel sheet
    # writer = pd.ExcelWriter(f'Rig List.xlsx', engine='openpyxl')
    # df = pd.DataFrame(jobs)
    # df.to_excel(writer, sheet_name='Jobs Processed', index=False)
    # writer.close()
    ######################################################################
    # Main Code- EDR Report Stuff below, API Configuration Stuff above
    ######################################################################


    # Variables

    # Define emojis as Unicode characters
    emoji_check = u'\u2705'  # ✅
    emoji_exclamation = u'\u2757'  # ❗
    emoji_x = u'\u274C'  # ❌

    attributeList = ['HookLoad', 'PumpPressure', 'BlockHeight', 'PumpSpm', 'PumpSpm2', 'RotaryTorque', 'BitPosition', 'SlipStatus']
    operators = ['Coterra', 'Upcurve', 'Alchemist', 'Endeavor', 'DB4', 'Piendeda', 'Surge', 'Black Swan', 'ConocoPhillips']
    processedJobList = []
    EDR_Operators = CFG['EDR_OperatorName']
    EDR_Rigs = CFG['EDR_Rigs']
    EDRJobs = []
    MR_Report_Ids = []
    MR_Report_Comments = []
    expectedAttr = CFG['ChannelsToOutput']
    attributeList = CFG['ChannelsToOutput']
    tmpJobs = []
    jobs = []
    jobsCurrTime = []
    jobsTimeBased = []
    count = 1
    jobcount = 1
    EDRchoice = CFG['ActiveEDRJobsOnly']
    jobsGetCurrTimeSecond = []
    testList = ['net_120413', 'net_164798', 'net_171957', 'net_176516', 'net_178067', 'net_178138', 'net_178269', 'net_178546', 'net_178570', 'net_178700', 'net_178781',
                'net_178797', 'net_178825', 'net_178869']
    report = ''
    well = ''
    wellList = []
    ZeroReportList = ['net_180486', 'net_180860']
    ReportErrorList = ['net_180006','net_178700','net_171957', 'net_181790', 'net_181159', 'net_180850', 'net_157732', 'net_158733', 'net_158799', 'net_158924', 'net_160986', 'net_163193', 'net_165070', 'net_166627', 'net_168850', 'net_169063', 'net_169683',
                       'net_169730', 'net_169890', 'net_170407', 'net_171655', 'net_171746', 'net_171874', 'net_172392', 'net_172515', 'net_172562', 'net_173338', 'net_174991',
                       'net_175348', 'net_175713', 'net_175967', 'net_175969', 'net_176105', 'net_176341', 'net_176508', 'net_176721', 'net_176750', 'net_177198', 'net_177229',
                       'net_177306', 'net_177803', 'net_177911', 'net_178030', 'net_178034', 'net_178060', 'net_178120', 'net_178126', 'net_178501', 'net_178600', 'net_178612',
                       'net_178738', 'net_178790', 'net_178816', 'net_178948', 'net_178971', 'net_179127', 'net_179163', 'net_179173', 'net_179181', 'net_179198', 'net_179201']

    # jobname = 'Janecka SA 4-H'
    # jobid = welldataAPI.getJobsbyWellName(token, CFG, jobname)
    # print(jobid)
    # #looking for wellname, put in job ID as placeholder, doesn't matter what jobid
    # jobid2 = welldataAPI.getJobs(URLs_v1['getJobs'], token, CFG, take=1000,jobStatus="ActiveJobs", wellname= jobname, jobid='net_177741')
    # print(jobid2)



    # # Get Jobs based on EDR vs All
    # if CFG['ActiveEDRJobsOnly'] == 0:
    #     # Getting Jobs -> will list through operator to retrieve rigs/
    #     for w in EDR_Operators:
    #         temp = welldataAPI.getJobs(URLs_v1['getJobs'], token, CFG, take=1000,jobStatus="ActiveJobs", operator=w)  # current URL Works for Getwells, testing below
    #         tmpJobs = tmpJobs + temp
    #
    # else:
    #     tmpJobs = welldataAPI.getJobs(URLs_v1['getJobs'], token, CFG, take=1000, total=False, jobStatus="ActiveJobs")  # current URL Works for Getwells, testing below

    #post for Curr Information of attributes


    # Testing
    lookup_table = {}
    tmpAllJobs = welldataAPI.getJobs(URLs_v1['getJobs'], token, CFG, take=1000, total=False, jobStatus="ActiveJobs")
    for w in tmpAllJobs:
        #jobs.append([w['id'], w['assetInfoList'][0]['owner'], w['assetInfoList'][0]['name']])
        key = f"{w['assetInfoList'][0]['owner']} {w['assetInfoList'][0]['name']}"
        # key = w['siteInfoList'][0]['owner']
        value = w['id']
        lookup_table[key] = value

    # Creating our Skinny table
    # Print the lookup table
    print("Lookup Table:")
    for key, value in lookup_table.items():
        print(f"Key: {key} | Value: {value}")


    for w in EDR_Rigs:
        for key, value in lookup_table.items():
            if w in key:
                tmpJobs.append(value)
    print("Temp Jobs:")
    for t in tmpJobs:
        print(t)

    try:
        for w in tmpJobs:
            well = str(w)
            attsLst = []
            # if jobcount % 80 == 0:
            #     token = welldataAPI.getToken(CFG['APIUrl'], CFG['appID'], CFG['username'], CFG['password'])
            # if jobcount % 20 == 0:
            #     time.sleep(2)
            # print(f'processing job {jobcount} of {len(tmpJobs)} with job id: {w["id"]}')
            # if well in ReportErrorList:
            #     jobcount = jobcount + 1
            #     continue

            # variables:
            holder = []
            HookLoadbool = emoji_x
            PumpPressurebool = emoji_x
            BlockHeightbool = emoji_x
            PumpSpmbool = emoji_x
            PumpSpm2bool = emoji_x
            PumpSpm3bool = emoji_x
            RotaryTorquebool = emoji_x
            BitPositionbool = emoji_x
            BitStatusbool = emoji_x
            SlipStatusbool = emoji_x
            tpDriveRPM = emoji_x
            comment = 'NA'
            comment24 = 'NA'
            reportDate = ''
            reportID = ''
            reportStatus = ''
            realTime = emoji_x
            tpDriveTorq = emoji_x
            weightonBit = emoji_x
            RP_Fast = emoji_x
            tHookLoad = emoji_x



            # Attribute Values
            HookLoad_val = ''
            PumpPressure_val = ''
            BlockHeight_val = ''
            PumpSpm_val = ''
            PumpSpm2_val = ''
            PumpSpm3_val = ''
            RotaryTorque_val = ''
            tpDriveRPM_val = ''
            tpDriveTorq_val = ''
            weightonBit_val = ''
            BitPosition_val = ''
            BitStatus_val = ''
            RP_Fast_val = ''
            SlipStatus_val = ''
            tHookLoad_val = ''


            # Checking for real time data capability
            rTime = welldataAPI.getApiCall(URLs_v1['getJobsIdCapabilities'], token, CFG, jobId=well)
            if 'realTime' in str(rTime):
                if rTime[0]['realTime'] == 'Supported':
                    realTime = emoji_check


            # 2 Get Attribute for job
            q = welldataAPI.getApiCall(URLs_v1['getAttributes'], token, CFG, jobId=well)


            # append attribute fields to holder
            for c in q[0]['attributes']:
                if len(c) == 0:
                    continue
                # print(f'processing job {count} of {len(q[0]["attributes"])} for jobid: {w["id"]} with ID: {c["id"]} and has data equal to : {c["hasData"]}')
                count = count + 1
                if c['hasData'] == True and c["id"] == 'HookLoad':
                    #HookLoadbool = emoji_check
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)
                if c['hasData'] == True and c["id"] == 'PumpPressure':
                    #PumpPressurebool = emoji_check
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)
                if c['hasData'] == True and c["id"] == 'BlockHeight':
                    #BlockHeightbool = emoji_check
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)
                if c['hasData'] == True and c["id"] == 'PumpSpm':
                    #PumpSpmbool = emoji_check
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)
                if c['hasData'] == True and c["id"] == 'PumpSpm2':
                    #PumpSpm2bool = emoji_check
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)
                if c['hasData'] == True and c["id"] == 'PumpSpm3':
                    #PumpSpm3bool = emoji_check
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)
                if c['hasData'] == True and c["id"] == 'RotaryTorque':
                    #RotaryTorquebool = emoji_check
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)
                if c['hasData'] == True and c["id"] == 'TopDrvRpm':  # tpDriveRPM
                    #tpDriveRPM = emoji_check
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)
                if c['hasData'] == True and c["id"] == 'TopDrvTorque':  # tpDriveTorq
                    #tpDriveTorq = emoji_check
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)
                if c['hasData'] == True and c["id"] == 'BitWeight':  # WOB
                    #weightonBit = emoji_check
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)
                if c['hasData'] == True and c["id"] == 'BitPosition':  # BitPosition
                    #BitPositionbool = emoji_check
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)
                if c['hasData'] == True and c["id"] == 'BitStatus':  # BitStatus
                    #BitStatusbool = emoji_check
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)
                if c['hasData'] == True and c["id"] == 'FastRopFtHr':  # ROP-F
                    #RP_Fast = emoji_check
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)
                if c['hasData'] == True and c["id"] == 'SlipStatus':  # SlipStatus
                    #SlipStatusbool = emoji_check
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)
                if c['hasData'] == True and c["id"] == 'TrigHkld':  # T-HL
                    #tHookLoad = emoji_check
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)

            if len(attsLst)== 0:
                continue
            to_time = datetime_string_to
            from_time = datetime_string_from
            formatted_to_time = datetime.fromisoformat(to_time)
            formatted_from_time = datetime.fromisoformat(from_time)
            hist_interval = CFG['HistoricInterval']
            hist_payload = welldataAPI.HistoricalTimeRequest(attributes=attsLst, toTime=to_time, fromTime=from_time, interval=hist_interval)
            hist = welldataAPI.historical_data_time(well, hist_payload.json(exclude_unset=True), token=token)
            jobsTimeBased.append([well, hist])

            attribute_mapping = {}

            if len(hist['timeRecords']) ==  0:
                continue
            values_timestamp_0 = hist['timeRecords'][0]['values']

            # Iterate over the 'attributes' list and map each attribute to its value from timestamp 0
            for i, attribute in enumerate(hist['attributes']):
                attribute_id = attribute['id']
                attribute_value = values_timestamp_0[i][1]
                attribute_mapping[attribute_id] = attribute_value

            print(attribute_mapping)

            # Iterate over the items in the dictionary
            for key, value in attribute_mapping.items():
                if key == 'HookLoad':
                    HookLoad_val = value
                    if isinstance(HookLoad_val, float):
                        if thresholdCheck(CFG['HookLoadbool_min'], CFG['HookLoadbool_max'], float(HookLoad_val)) == False:
                            HookLoadbool = emoji_exclamation
                        else:
                            HookLoadbool = emoji_check

                if key == 'PumpPressure':
                    PumpPressure_val = value
                    if isinstance(PumpPressure_val, float):
                        if thresholdCheck(CFG['PumpPressurebool_min'], CFG['PumpPressurebool_max'], float(PumpPressure_val)) == False:
                            PumpPressurebool = emoji_exclamation
                        else:
                            PumpPressurebool = emoji_check

                if key == 'BlockHeight':
                    BlockHeight_val = value
                    if isinstance(BlockHeight_val, float):
                        if thresholdCheck(CFG['BlockHeightbool_min'], CFG['BlockHeightbool_max'], float(BlockHeight_val)) == False:
                            BlockHeightbool = emoji_exclamation
                        else:
                            BlockHeightbool = emoji_check
                if key == 'PumpSpm':
                    PumpSpm_val = value
                    if isinstance(PumpSpm_val, float):
                        if thresholdCheck(CFG['PumpSpmbool_min'], CFG['PumpSpmbool_max'], float(PumpSpm_val)) == False:
                            PumpSpmbool = emoji_exclamation
                        else:
                            PumpSpmbool = emoji_check

                if key == 'PumpSpm2':
                    PumpSpm2_val = value
                    if isinstance(PumpSpm2_val, float):
                        if thresholdCheck(CFG['PumpSpm2bool_min'], CFG['PumpSpm2bool_max'], float(PumpSpm2_val)) == False:
                            PumpSpm2bool = emoji_exclamation
                        else:
                            PumpSpm2bool = emoji_check

                if key == 'PumpSpm3':
                    PumpSpm3_val = value
                    if isinstance(PumpSpm3_val, float):
                        if thresholdCheck(CFG['PumpSpm3bool_min'], CFG['PumpSpm3bool_max'], float(PumpSpm3_val)) == False:
                            PumpSpm3bool = emoji_exclamation
                        else:
                            PumpSpm3bool = emoji_check

                if key == 'RotaryTorque':
                    RotaryTorque_val = value
                    if isinstance(RotaryTorque_val, float):
                        if thresholdCheck(CFG['RotaryTorquebool_min'], CFG['RotaryTorquebool_max'], float(RotaryTorque_val)) == False:
                            RotaryTorquebool = emoji_exclamation
                        else:
                            RotaryTorquebool = emoji_check

                if key == 'TopDrvRpm':  # tpDriveRPM
                    tpDriveRPM_val = value
                    if isinstance(tpDriveRPM_val, float):
                        if thresholdCheck(CFG['tpDriveRPM_min'], CFG['tpDriveRPM_max'], float(tpDriveRPM_val)) == False:
                            tpDriveRPM = emoji_exclamation
                        else:
                            tpDriveRPM = emoji_check

                if key == 'TopDrvTorque':  # tpDriveTorq
                    tpDriveTorq_val = value
                    if isinstance(tpDriveTorq_val, float):
                        if thresholdCheck(CFG['tpDriveTorq_min'], CFG['tpDriveTorq_max'], float(tpDriveTorq_val)) == False:
                            tpDriveTorq = emoji_exclamation
                        else:
                            tpDriveTorq = emoji_check

                if key == 'BitWeight':  # WOB
                    weightonBit_val = value
                    if isinstance(tpDriveTorq_val, float):
                        if thresholdCheck(CFG['WOB_min'], CFG['WOB_max'], float(weightonBit_val)) == False:
                            weightonBit = emoji_exclamation
                        else:
                            weightonBit = emoji_check

                if key == 'BitPosition':  # BitPosition
                    BitPosition_val = value
                    if isinstance(BitPosition_val, float):
                        if thresholdCheck(CFG['BitPositionbool_min'], CFG['BitPositionbool_max'], float(BitPosition_val)) == False:
                            BitPositionbool = emoji_exclamation
                        else:
                            BitPositionbool = emoji_check

                if key == 'BitStatus':  # BitStatus
                    BitStatus_val = value
                    if isinstance(BitStatus_val, float):
                        if thresholdCheck(CFG['BitStatusbool_min'], CFG['BitStatusbool_max'], float(BitStatus_val)) == False:
                            BitStatusbool = emoji_exclamation
                        else:
                            BitStatusbool = emoji_check

                if key == 'FastRopFtHr':  # ROP-F
                    RP_Fast_val = value
                    if isinstance(RP_Fast_val, float):
                        if thresholdCheck(CFG['RP_Fast_min'], CFG['RP_Fast_max'], float(RP_Fast_val)) == False:
                            RP_Fast = emoji_exclamation
                        else:
                            RP_Fast = emoji_check

                if key == 'SlipStatus':  # SlipStatus
                    SlipStatus_val = value
                    if isinstance(SlipStatus_val, float):
                        if thresholdCheck(CFG['SlipStatusbool_min'], CFG['SlipStatusbool_max'], float(SlipStatus_val)) == False:
                            SlipStatusbool = emoji_exclamation
                        else:
                            SlipStatusbool = emoji_check

                if key == 'TrigHkld':  # T-HL
                    tHookLoad_val = value
                    if isinstance(tHookLoad_val, float):
                        if thresholdCheck(CFG['tHookLoad_min'], CFG['tHookLoad_max'], float(tHookLoad_val)) == False:
                            tHookLoad = emoji_exclamation
                        else:
                            tHookLoad = emoji_check









            # getting the comment portion


            r = welldataAPI.getReports(URLs_v1['getReportsClassificationReportGroup'], token, CFG, jobId=well, reportGroupId=2, classification='daily')
            MR_Report_Ids.append([well, r])
            # Checking iif the morning reports exists for job ID
            if well in ZeroReportList:
                continue
            elif len(r) == 0:
                ZeroReportList.append(well)
                # for z in ZeroReportList:
                #     print(z)
                continue


            else:
                # Go through Id's and pull pdf downloads of reports
                report = welldataAPI.getReports(URLs_v1['getReportsClassificationReportGroupFileFormat'], token, CFG, jobId=well, reportGroupId=2, classification='daily',fileFormat='JSON', reportId=r[0]['id'])

                if 'GenericAmericanMorningReportDW' in str(report):
                    # continue
                    reportDate = report[0]['Reports'][0]['GenericAmericanMorningReportDW']['Header']['Date']
                    comment = report[0]['Reports'][0]['GenericAmericanMorningReportDW']['Header']['OpsAtReportTime']
                    if 'OpsNext24' in str(report):
                        comment24 = report[0]['Reports'][0]['GenericAmericanMorningReportDW']['Header']['OpsNext24']
                    reportID = report[0]['Reports'][0]['GenericAmericanMorningReportDW']['ReportAttributes']['ReportID']
                    reportStatus = report[0]['Reports'][0]['GenericAmericanMorningReportDW']['ReportAttributes']['ReportStatus']
                    # Getting the Activity code and comments
                    for rep in report[0]['Reports'][0]['GenericAmericanMorningReportDW']['ActivityDetails']['Items']:
                        if 'ActCode' in rep:
                            MR_Report_Comments.append([well, r[0]['id'], r[0]['date'], rep['ActCode'], rep['DescriptionOfWork']])
                        else:
                            continue
                    # 'HandPMorningReport'
                elif 'HandPMorningReport' in str(report):
                    reportDate = report[0]['Reports'][0]['HandPMorningReport']['Header']['Date']
                    comment = report[0]['Reports'][0]['HandPMorningReport']['Operations']['PresentOp']
                    comment24 = ''
                    reportID = report[0]['Reports'][0]['HandPMorningReport']['ReportAttributes']['ReportID']
                    reportStatus = report[0]['Reports'][0]['HandPMorningReport']['ReportAttributes']['ReportStatus']
                    # Getting the Activity code and comments
                    for rep in report[0]['Reports'][0]['HandPMorningReport']['TimeSummary']['Items']:
                        if 'ActivityCode' in rep:
                            MR_Report_Comments.append(
                                [well, r[0]['id'], r[0]['date'], rep['ActivityCode'], rep['ActivityDetails']])
                        else:
                            continue
                    # 'ScanMorningReport'
                elif 'ScanMorningReport' in str(report):
                    reportDate = report[0]['Reports'][0]['ScanMorningReport']['Header']['Date']
                    reportID = report[0]['Reports'][0]['ScanMorningReport']['ReportAttributes']['ReportID']
                    reportStatus = report[0]['Reports'][0]['ScanMorningReport']['ReportAttributes']['ReportStatus']
                    comment = report[0]['Reports'][0]['ScanMorningReport']['Header']['PresentOp']
                    # Getting the Activity code and comments
                    for rep in report[0]['Reports'][0]['ScanMorningReport']['TimeBreakDown']['Items']:
                        if 'ActivityCode' in rep:
                            MR_Report_Comments.append(
                                [well, r[0]['id'], r[0]['date'], rep['ActivityCode'], rep['ActivityDetails']])
                        else:
                            continue

                    # 'RapadMorningReport'
                elif 'RapadMorningReport' in str(report):
                    reportDate = report[0]['Reports'][0]['RapadMorningReport']['Header']['ReportDate']
                    comment = report[0]['Reports'][0]['RapadMorningReport']['Header']['OperationsActivityCurrent']
                    comment24 = report[0]['Reports'][0]['RapadMorningReport']['Header']['OperationsActivityNext24Hours']
                    reportID = report[0]['Reports'][0]['RapadMorningReport']['ReportAttributes']['ReportID']
                    reportStatus = report[0]['Reports'][0]['RapadMorningReport']['ReportAttributes']['ReportStatus']

                    # Getting the Activity code and comments
                    for rep in report[0]['Reports'][0]['RapadMorningReport']['ActivityDetails']['Items']:
                        if 'OperationsActivityCode' in rep:
                            MR_Report_Comments.append([well, r[0]['id'], r[0]['date'], rep['OperationsActivityCode'], rep['OperationsActivityDescription']])
                        else:
                            continue
                    # 'PattersonMorningReportRevB'
                elif 'PattersonMorningReportRevB' in str(report):
                    reportDate = report[0]['Reports'][0]['PattersonMorningReportRevB']['Header']['ReportDate']
                    comment = report[0]['Reports'][0]['PattersonMorningReportRevB']['OperationsCasingDetails']['operations_at_report_time']
                    if 'operations_next_24_hours' in report[0]['Reports'][0]['PattersonMorningReportRevB']['OperationsCasingDetails']:
                        comment24 = report[0]['Reports'][0]['PattersonMorningReportRevB']['OperationsCasingDetails']['operations_next_24_hours']
                    reportID = report[0]['Reports'][0]['PattersonMorningReportRevB']['ReportAttributes']['ReportID']
                    reportStatus = report[0]['Reports'][0]['PattersonMorningReportRevB']['ReportAttributes']['ReportStatus']

                    # Getting the Activity code and comments
                    for rep in report[0]['Reports'][0]['PattersonMorningReportRevB']['ActivityDetails']['Items']:
                        if 'details' in rep:
                            if 'code' in rep:
                                code = rep['code']
                            else:
                                code = ''
                            MR_Report_Comments.append(
                                [well, r[0]['id'], r[0]['date'], code, rep['details']])
                        else:
                            continue
                else:
                    processedJobList.append([well, str(report[0]['Reports'][0])])
                    # print(processedJobList)
                    continue


            # Appending jobs to dataFrame
            job = welldataAPI.getJobs(URLs_v1['getJobsId'], token, CFG, take=1000, total=False, jobStatus="ActiveJobs", jobId=well)
            # Appending for EDRJobs
            holder.append(f'{job[0]["assetInfoList"][0]["owner"]} {job[0]["assetInfoList"][0]["name"]}')
            holder.append(job[0]['siteInfoList'][0]['owner'])
            holder.append(job[0]['name'])
            #holder.append(well)
            holder.append(realTime)
            holder.append(str(HookLoadbool))
            holder.append(str(PumpPressurebool))
            holder.append(str(BlockHeightbool))
            holder.append(str(PumpSpmbool))
            holder.append(str(PumpSpm2bool))
            holder.append(str(PumpSpm3bool))
            holder.append(str(RotaryTorquebool))
            holder.append(str(tpDriveRPM))
            holder.append(str(tpDriveTorq))
            holder.append(str(weightonBit))
            holder.append(str(BitPositionbool))
            holder.append(str(BitStatusbool))
            holder.append(str(RP_Fast))
            holder.append(str(SlipStatusbool))
            holder.append(str(tHookLoad))
            holder.append(comment)
            holder.append(comment24)
            holder.append(reportID)
            holder.append(reportDate)
            # holder.append(reportStatus)
            count = count + 1
            jobcount = jobcount + 1

            EDRJobs.append(holder)

    except Exception as ex:
        wellList.append(well)
        print(well)
        print(report)
        logging.error("Error sending request to server")
        logging.error(f"Exception: {ex}")
        pass

    # Writing DataFrame to Excel sheet
    writer = pd.ExcelWriter(f'EDRJobsListTest {today} .xlsx', engine='openpyxl')
    header = ['Rig', 'Operator', 'Well name', 'Real Time', 'HookLoad', 'PumpPressure', 'BlockHeight', 'PumpSpm', 'PumpSpm2', 'PumpSpm3', 'RotaryTorque', 'TopDrive RPM',
              'TopDrive Torque', 'WOB', 'BitPosition', 'BitStatus', 'ROP-F', 'SlipStatus', 'T-HL', 'Comments', 'Next 24 Hr Comments', 'Report Id', 'Report Date']
    df = pd.DataFrame(tmpJobs)
    df.to_excel(writer, sheet_name='Jobs Processed', index=False)

    df = pd.DataFrame(EDRJobs)
    df = df.sort_values(1, ascending=False)
    df.to_excel(writer, sheet_name='EDR Report', index=False, header = header)  # header = header

    # df = pd.DataFrame(MR_Report_Ids)
    # df.to_excel(writer, sheet_name='Report_Ids', index=False, header=['Job Ids', 'Report IDs'])

    header2 = ['JobID', 'Report Id', 'Report Date', 'Activity Code', 'Details of Operation']

    df = pd.DataFrame(MR_Report_Comments)
    df.to_excel(writer, sheet_name='Report_Comment', index=False)  # , header = header2

    df = pd.DataFrame(processedJobList)
    df.to_excel(writer, sheet_name='Jobs not found', index=False, )

    df = pd.DataFrame(ZeroReportList)
    df.to_excel(writer, sheet_name='No Report List', index=False, )

    df = pd.DataFrame(jobsTimeBased)
    df.to_excel(writer, sheet_name='Time Based Pull', index=False, )


    # List of sheet names to hide
    sheets_to_hide = ['Jobs Processed',  'Jobs not found', 'No Report List'] #'Report_Ids',

    # Loop through the list of sheets to hide
    for sheet_name in sheets_to_hide:
        # Get the sheet object
        sheet_to_hide = writer.sheets[sheet_name]
        # Hide the sheet
        sheet_to_hide.sheet_state = 'hidden'
    writer.close()

    wb = openpyxl.load_workbook(f'EDRJobsListTest {today} .xlsx')
    for worksheet in wb.worksheets:
        print(worksheet.title)

    # Define the conditional formatting

    # Select the "EDR Jobs" worksheet
    ws = wb['EDR Report']

    # Define the format objects
    from openpyxl.formatting.rule import IconSet, FormatObject, ColorScale, IconSetRule, CellIsRule, Rule
    from openpyxl.styles import PatternFill, Font
    # red = '#FFFF0000'
    # yellow = '#FFFFFF00'
    # green = '#FF00FF00'
    # rule = ColorScaleRule(start_type='min', start_color='FFFF0000',
    #                       mid_type='percentile', mid_value=50, mid_color='FFFFFF00',
    #                       end_type='max', end_color='FF00FF00')

    red_fill = PatternFill(start_color='FFFF0000',
                           end_color='FFFF0000',
                           fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFFFF00',
                              end_color='FFFFFF00',
                              fill_type='solid')
    green_fill = PatternFill(start_color='FF00FF00',
                             end_color='FF00FF00',
                             fill_type='solid')

    red_font = Font(color='FFFFFFFF', bold=True)
    yellow_font = Font(color='FF000000', bold=True)
    green_font = Font(color='FFFFFFFF', bold=True)

    icon_rule = IconSetRule('3TrafficLights1', 'num', [0, 33, 67])
    icon_rule.dxfId = 0

    # Set the worksheet to display the icon set

    ws.conditional_formatting.rule = Rule(type="iconSet", priority=1)
    ws.conditional_formatting.rule.operator = "lessThan"
    ws.conditional_formatting.rule.formula = ["0"]

    ws.conditional_formatting.add('D1:S150', icon_rule)
    wb.close()

    # Sending Email
    from EmailModule import send_email
    recipients = ['elvis.segbeaya@nov.com']
    for recipient in recipients:
        email_body = f'Hi {str(recipient)},\n \nPlease see the attached email for today"s EDR Data\n\n\nAlso, this is a test email.\n There"s no report Date, so I added the report completion status.'
        # send email
        send_email('Test Email From Elvis ', email_body, recipient, f'EDRJobsListTest {today} .xlsx')

    quit()


if __name__ == "__main__":
    main()
